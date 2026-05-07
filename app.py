import io

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from leadgen import campaign_composer_ui
from leadgen.backend import AirtableBackend
from leadgen.config import get_secrets
from leadgen.credits import display_credit_dashboard
from leadgen.dashboard import StatusDashboard
from leadgen.runner import execute_with_credit_tracking
from leadgen.sync_manager import sync_pending_leads, sync_with_verification


# --- CONSTANTS & CONFIGURATION ---
AIRTABLE_LEADS_TABLE = "tblKrC9hOxCuMMyZT"
AIRTABLE_LOG_TABLE = "log"
SCRAPPING_TOOL_ID = "maps_apify_apollo"
MAX_SYNC_PER_RUN = 5000

DEFAULT_INDUSTRIES = [
    "Dentist",
    "Family lawyer",
    "Immigration lawyer",
    "Interior Design",
    "Personal injury lawyer",
    "Property management",
    "Real Estate and Trust lawyer",
    "Med Spa",
    "Clinic Services",
    "Elder and Disabled Care",
    "Hotels and Leisure",
    "Restaurants and Bars",
    "Consulting Services (B2B)",
]


st.set_page_config(page_title="Lead Generation Engine", page_icon="🚀", layout="wide")


def _init_backend(secrets: dict, mode: str):
    """Initialize the selected backend. Returns a DataBackend instance."""
    if mode == "supabase":
        from leadgen.supabase_utils import SupabaseBackend
        return SupabaseBackend(secrets)
    else:
        return AirtableBackend(secrets, AIRTABLE_LEADS_TABLE, AIRTABLE_LOG_TABLE)


def main():
    secrets = get_secrets()

    # ── Backend toggle (Step 3.3) ──
    config_default = secrets.get("data_backend", "airtable")
    supabase_available = bool(secrets.get("supabase_db_url"))

    st.sidebar.header("⚙️ Data Backend")
    backend_options = ["airtable", "supabase"]
    default_index = backend_options.index(config_default) if config_default in backend_options else 0

    active_mode = st.sidebar.radio(
        "Storage backend",
        options=backend_options,
        index=default_index,
        format_func=lambda x: "Airtable (Legacy)" if x == "airtable" else "Supabase",
        disabled=False,
        key="backend_mode",
        help="Switch between Airtable and Supabase for lead storage.",
    )

    # Prevent Supabase if credentials missing
    if active_mode == "supabase" and not supabase_available:
        st.sidebar.error("Supabase credentials not configured. Falling back to Airtable.")
        active_mode = "airtable"

    # Visual indicator
    if active_mode == "supabase":
        st.sidebar.success("🟢 **SUPABASE MODE**")
    else:
        st.sidebar.warning("🟠 **AIRTABLE MODE (Legacy)**")

    if active_mode != config_default:
        st.sidebar.caption(f"⚠️ Override — config default is **{config_default}**")
    else:
        st.sidebar.caption("(config default)")

    st.sidebar.divider()

    # ── Init backend ──
    st.sidebar.header("🔌 Connection Status")
    try:
        backend = _init_backend(secrets, active_mode)
        if active_mode == "supabase":
            st.sidebar.success("Supabase Connected")
        else:
            st.sidebar.success("Airtable Connected")
    except Exception as e:
        st.sidebar.error(f"Backend Connection Failed: {e}")
        st.stop()

    if secrets["apify_token"]:
        st.sidebar.success("Apify Token Found")
    if secrets["apollo_key"]:
        st.sidebar.success("Apollo Key Found")
    if secrets["instantly_key"]:
        st.sidebar.success("Instantly Key Found")
    else:
        st.sidebar.warning("Instantly Key Missing")

    if secrets.get("gemini_key"):
        st.sidebar.success("Gemini Key Found")
    else:
        st.sidebar.warning("Gemini Key Missing (split disabled)")

    if secrets.get("millionverifier_key"):
        st.sidebar.success("MillionVerifier Key Found")
    else:
        st.sidebar.warning("MillionVerifier Key Missing (email verification disabled)")

    debug_mode = st.sidebar.checkbox("🔍 Debug Mode", value=False, help="Show API response details")
    st.session_state["debug_mode"] = debug_mode

    use_gemini_split_default = bool(secrets.get("gemini_key"))
    use_gemini_split = st.sidebar.checkbox(
        "🧩 Use Gemini Split (10 zones)",
        value=use_gemini_split_default,
        help="Splits the location into 10 sub-zones to increase Google Maps coverage. Falls back to single query on any Gemini error.",
        disabled=not bool(secrets.get("gemini_key")),
    )
    st.session_state["use_gemini_split"] = use_gemini_split

    display_credit_dashboard(
        secrets["apify_token"],
        secrets["apollo_key"],
        secrets["instantly_key"],
        debug=debug_mode,
    )

    if "industry_options" not in st.session_state:
        st.session_state["industry_options"] = backend.get_industry_options()

    st.title("🚀 Lead Generation Engine")

    # --- TABS ---
    show_legacy_generator = st.sidebar.checkbox("Show Legacy Generator", value=False, help="Show the Apify/Apollo scraping tab (legacy)")
    if show_legacy_generator:
        tab_gen, tab_man, tab_camp = st.tabs([
            "🚀 Lead Generator (Legacy)",
            "📝 Lead Manager",
            "📣 Campaign Composer",
        ])
    else:
        tab_man, tab_camp = st.tabs(["📝 Lead Manager", "📣 Campaign Composer"])
        tab_gen = None

    # --- TAB 1: GENERATOR (Legacy, hidden by default) ---
    if tab_gen is not None:
      with tab_gen:
        col1, col2, col3 = st.columns([2, 2, 1])
        with col1:
            def on_industry_change():
                st.session_state["search_query_input"] = st.session_state["industry_selector"]

            industry = st.selectbox(
                "Industry",
                st.session_state["industry_options"] or ["Generic"],
                key="industry_selector",
                on_change=on_industry_change
            )

            # Initialize search query if not present
            if "search_query_input" not in st.session_state:
                st.session_state["search_query_input"] = industry

            search_query = st.text_input("Search Query", key="search_query_input")

        with col2:
            city_input = st.text_input("City", placeholder="e.g. New York")
        with col3:
            max_leads = st.number_input("Max Leads", min_value=1, max_value=500, value=10, step=1)

        enrich_emails = st.checkbox("✨ Enrich with Emails? (Apollo)", value=True)

        if st.button("Find & Sync Leads", type="primary"):
            if not city_input:
                st.warning("Please enter a city.")
                # We cannot just return here because we are in a with block inside main
            else:
                status_dashboard = StatusDashboard()

                result_data, credit_used_apify, credit_used_apollo, credit_used_instantly = execute_with_credit_tracking(
                    secrets,
                    backend,
                    industry,
                    city_input,
                    max_leads,
                    enrich_emails,
                    status_dashboard,
                    leads_table_id=AIRTABLE_LEADS_TABLE,
                    scrapping_tool_id=SCRAPPING_TOOL_ID,
                    search_query=search_query,
                )

                status_dashboard.update_status("Execution Complete!", 100)

                backend.log_transaction(
                    industry=industry,
                    city_input=city_input,
                    total_scraped=result_data["total_scraped"],
                    new_added=result_data["new_added"],
                    enrich_used=enrich_emails,
                    status=result_data["status"],
                    error_msg=result_data["error_msg"],
                    credit_used_apify=credit_used_apify,
                    credit_used_apollo=credit_used_apollo,
                    credit_used_instantly=credit_used_instantly,
                    instantly_added=result_data.get("instantly_added"),
                    search_query=search_query,
                )

                if result_data["status"] in ("Success", "Zero Results"):
                    credit_info = []
                    if credit_used_apify is not None:
                        credit_info.append(f"Apify: ${credit_used_apify:.4f}")
                    if credit_used_apollo is not None:
                        credit_info.append(f"Apollo: {credit_used_apollo} credits")
                    credit_str = f" ({', '.join(credit_info)})" if credit_info else ""

                    st.success(f"✅ Scraped {result_data['total_scraped']}, Added {result_data['new_added']}.{credit_str}")

                    # Invalidate cached lead_df to force reload on next visit to Manager
                    if "lead_df" in st.session_state:
                        del st.session_state["lead_df"]

                    if result_data["new_records"]:
                        st.dataframe(pd.DataFrame(result_data["new_records"]))
                    else:
                        st.info("No new unique leads found.")
                else:
                    st.error(f"❌ Execution failed: {result_data['error_msg']}")

                with st.sidebar:
                    st.markdown("---")
                    st.caption("🔄 Refreshing credit dashboard...")
                display_credit_dashboard(
                    secrets["apify_token"],
                    secrets["apollo_key"],
                    secrets["instantly_key"],
                    debug=st.session_state.get("debug_mode", False),
                )

    # --- TAB 2: LEAD MANAGER ---
    with tab_man:
        st.info("💡 **Lead Manager**: Import leads via CSV, review pending syncs, and push to Instantly.")

        # ── Industry Management ──
        if "industries" not in st.session_state:
            st.session_state["industries"] = list(DEFAULT_INDUSTRIES)

        with st.expander("🏷️ Manage Industries", expanded=False):
            st.caption("Industries used for campaign targeting. Edit the list below.")
            col_add, col_remove = st.columns(2)
            with col_add:
                new_industry = st.text_input("Add industry", key="new_industry_input", placeholder="e.g. Veterinary Clinic")
                if st.button("➕ Add", key="add_industry_btn") and new_industry.strip():
                    name = new_industry.strip()
                    if name not in st.session_state["industries"]:
                        st.session_state["industries"].append(name)
                        st.session_state["industries"].sort()
                        st.rerun()
                    else:
                        st.warning(f"'{name}' already exists.")
            with col_remove:
                if st.session_state["industries"]:
                    to_remove = st.selectbox("Remove industry", st.session_state["industries"], key="remove_industry_select")
                    if st.button("🗑️ Remove", key="remove_industry_btn"):
                        st.session_state["industries"].remove(to_remove)
                        st.rerun()
            st.write(f"**{len(st.session_state['industries'])} industries**: {', '.join(st.session_state['industries'])}")

        # ── Import Leads ──
        with st.expander("📥 Import Leads", expanded=False):
            TEMPLATE_COLUMNS = [
                "company_name", "industry", "website", "city", "state",
                "postal_code", "postal_address", "phone", "rating",
                "contact_name", "contact_email", "contact_position",
                "competitor1", "competitor2", "competitor3",
            ]

            # Build Excel template with two sheets
            def build_template_xlsx() -> bytes:
                wb = Workbook()

                # Sheet 1: Leads template
                ws_leads = wb.active
                ws_leads.title = "Leads"
                for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, 1):
                    ws_leads.cell(row=1, column=col_idx, value=col_name)
                # Set column widths
                for col_idx in range(1, len(TEMPLATE_COLUMNS) + 1):
                    ws_leads.column_dimensions[ws_leads.cell(row=1, column=col_idx).column_letter].width = 18

                # Sheet 2: Industries reference
                ws_industries = wb.create_sheet("Industries")
                ws_industries.cell(row=1, column=1, value="industry")
                for row_idx, ind in enumerate(st.session_state["industries"], 2):
                    ws_industries.cell(row=row_idx, column=1, value=ind)
                ws_industries.column_dimensions["A"].width = 30

                # Data validation on industry column (column B = index 2)
                industry_count = len(st.session_state["industries"])
                if industry_count > 0:
                    dv = DataValidation(
                        type="list",
                        formula1=f"Industries!$A$2:$A${industry_count + 1}",
                        allow_blank=True,
                    )
                    dv.error = "Pick an industry from the Industries sheet."
                    dv.errorTitle = "Invalid Industry"
                    dv.prompt = "Select an industry from the dropdown."
                    dv.promptTitle = "Industry"
                    ws_leads.add_data_validation(dv)
                    dv.add(f"B2:B10000")  # Column B = industry

                buf = io.BytesIO()
                wb.save(buf)
                return buf.getvalue()

            st.download_button(
                "📄 Download Import Template (.xlsx)",
                data=build_template_xlsx(),
                file_name="vyzz_lead_import_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Excel template with lead columns + industry dropdown. Fill Sheet 1 and upload below.",
            )

            st.caption("**Required**: `company_name`. **Recommended**: `contact_email`, `industry`, `website`.")
            st.caption("The `industry` column has a dropdown — pick from the Industries sheet.")

            # File upload (CSV or Excel)
            uploaded_file = st.file_uploader(
                "Upload filled template",
                type=["csv", "xlsx"],
                key="lead_uploader",
                help="Upload the filled template (CSV or Excel).",
            )

            if uploaded_file is not None:
                try:
                    if uploaded_file.name.endswith(".xlsx"):
                        import_df = pd.read_excel(uploaded_file, sheet_name="Leads", engine="openpyxl")
                    else:
                        import_df = pd.read_csv(uploaded_file)

                    # Validate
                    if "company_name" not in import_df.columns:
                        st.error("File must have a `company_name` column.")
                    elif import_df.empty:
                        st.warning("File is empty.")
                    else:
                        # Drop columns not in template
                        valid_cols = [c for c in import_df.columns if c in TEMPLATE_COLUMNS]
                        import_df = import_df[valid_cols]

                        # Drop rows with no company_name
                        import_df = import_df.dropna(subset=["company_name"])
                        import_df = import_df[import_df["company_name"].astype(str).str.strip() != ""]

                        # Warn on unknown industries
                        if "industry" in import_df.columns:
                            known = set(st.session_state["industries"])
                            unknown = set(import_df["industry"].dropna().unique()) - known
                            if unknown:
                                st.warning(f"⚠️ Unknown industries: {', '.join(sorted(unknown))}. They will still be imported.")

                        st.write(f"**{len(import_df)} valid leads** found.")
                        st.dataframe(import_df.head(10), use_container_width=True)

                        # Batch metadata from most common values
                        batch_industry = import_df["industry"].mode().iloc[0] if "industry" in import_df.columns and not import_df["industry"].dropna().empty else "Unknown"
                        batch_city = import_df["city"].mode().iloc[0] if "city" in import_df.columns and not import_df["city"].dropna().empty else "Unknown"

                        if st.button(f"🚀 Import {len(import_df)} Leads", type="primary", key="import_btn"):
                            records = import_df.where(import_df.notna(), None).to_dict("records")
                            with st.spinner(f"Importing {len(records)} leads..."):
                                batch_id = backend.batch_create(
                                    records,
                                    source_tool="csv_import",
                                    industry=batch_industry,
                                    city=batch_city,
                                )
                            if batch_id is not None:
                                st.success(f"✅ Imported {len(records)} leads (batch: `{batch_id}`)")
                                if "lead_df" in st.session_state:
                                    del st.session_state["lead_df"]
                            else:
                                if active_mode == "airtable":
                                    st.success(f"✅ Imported {len(records)} leads to Airtable")
                                    if "lead_df" in st.session_state:
                                        del st.session_state["lead_df"]
                                else:
                                    st.error("Import failed. Check the logs above.")
                except Exception as e:
                    st.error(f"Failed to read file: {e}")

        st.divider()

        if st.button("🔄 Refresh Data"):
            if "lead_df" in st.session_state:
                del st.session_state["lead_df"]
            st.rerun()

        if "lead_df" not in st.session_state:
            source_label = "Supabase" if active_mode == "supabase" else "Airtable"
            with st.spinner(f"Fetching leads from {source_label}..."):
                raw_leads = backend.fetch_all_leads()
                st.session_state["lead_df"] = pd.DataFrame(raw_leads)

        # Base DataFrame
        df = st.session_state["lead_df"].copy()
        
        # Ensure timestamp columns exist
        if "last_modified_at" not in df.columns:
            df["last_modified_at"] = None
        if "last_synced_at" not in df.columns:
            df["last_synced_at"] = None
            
        # Convert to datetime for comparison
        df["last_modified_at"] = pd.to_datetime(df["last_modified_at"], errors='coerce', utc=True)
        df["last_synced_at"] = pd.to_datetime(df["last_synced_at"], errors='coerce', utc=True)

        # --- FILTER PENDING UPDATES ---
        # Logic: last_modified > last_synced OR last_synced is NaT (Never synced)
        # Note: We must handle NaT carefully.
        # If synced is NaT -> Pending.
        # If modified is NaT (shouldn't happen for valid records) -> Not pending? Or Pending? Assume Modified exists.
        
        def is_pending(row):
            mod = row["last_modified_at"]
            syn = row["last_synced_at"]

            # Historical backfill guard: rows explicitly marked as having a
            # stale Instantly ID (Phase D Step 3.8 C2a-stale cohort) must NOT
            # be auto-re-pushed. Their instantly_lead_id was cleared on
            # purpose, and auto-pushing them re-burns MillionVerifier credits
            # and risks creating duplicates if the original lead still exists
            # elsewhere in Instantly. Operator must re-verify cohort manually.
            status = row.get("instantly_statuts")
            if status == "stale_reference_cleared":
                return False

            if pd.isna(mod): return False # Should not happen if Airtable tracks it
            if pd.isna(syn): return True # Never synced

            # Recovery: leads wrongly blocked as catch_all/unknown need re-syncing.
            # They have last_synced_at set but were never actually added to Instantly.
            v_status = row.get("verification_status", "")
            if status == "Blocked" and str(v_status).strip().lower() in ("catch_all", "unknown", ""):
                return True

            return mod > syn

        # Apply pending filter
        # We need a mask
        pending_mask = df.apply(is_pending, axis=1)
        pending_df = df[pending_mask]
        
        pending_count = len(pending_df)
        if pending_count > MAX_SYNC_PER_RUN:
            st.warning(
                f"Sync is capped at {MAX_SYNC_PER_RUN} leads per run. "
                f"{pending_count - MAX_SYNC_PER_RUN} will remain pending for the next refresh."
            )
        
        # --- Industry filter for focused sync ---
        available_industries = sorted(
            pending_df["industry"].dropna().unique().tolist()
        ) if "industry" in pending_df.columns and not pending_df.empty else []

        col_header, col_filter = st.columns([2, 2])
        with col_header:
            st.subheader(f"⏳ Pending Updates: {pending_count} Leads")
        with col_filter:
            selected_industry = st.selectbox(
                "Filter by industry",
                options=["All Industries"] + available_industries,
                key="sync_industry_filter",
            )

        if selected_industry != "All Industries":
            pending_df = pending_df[pending_df["industry"] == selected_industry]
            pending_count = len(pending_df)
            st.info(f"Filtered to **{pending_count}** leads in **{selected_industry}**")

        col_sync, _ = st.columns([1, 3])
        with col_sync:
             # Sync Button (Active only if pending items exist)
             if st.button(f"🚀 Sync {pending_count} Updates", type="primary", disabled=pending_count == 0):
                 with st.status("Syncing to Instantly...", expanded=True) as status:
                     pending_records = pending_df.to_dict("records")

                     if secrets.get("millionverifier_key"):
                         # Smart Verification & Delta Cleanup workflow
                         sync_result = sync_with_verification(
                             pending_records,
                             backend,
                             secrets=secrets,
                             debug_mode=debug_mode,
                             max_records=MAX_SYNC_PER_RUN,
                             status=status,
                         )
                     else:
                         # Fallback: no verification, sync directly
                         st.warning("⚠️ MillionVerifier key missing – syncing without email verification.")
                         sync_result = sync_pending_leads(
                             pending_records,
                             backend,
                             secrets=secrets,
                             debug_mode=debug_mode,
                             max_records=MAX_SYNC_PER_RUN,
                             status=status,
                         )

                     if sync_result.get("error"):
                         st.error("Failed to update Airtable labels.")
                         st.stop()

                     # Save results to session state for persistent logging
                     st.session_state["last_sync_results"] = sync_result

                     # Refresh
                     del st.session_state["lead_df"]
                     st.rerun()

        # Display Pending Table (Read-Only)
        if pending_count > 0:
            st.dataframe(
                pending_df,
                use_container_width=True,
                column_config={
                    "last_modified_at": st.column_config.DatetimeColumn("Modified", format="D MMM HH:mm"),
                    "last_synced_at": st.column_config.DatetimeColumn("Last Synced", format="D MMM HH:mm"),
                    "id": None,
                    "createdTime": None
                },
                column_order=[
                    "company_name", 
                    "industry", 
                    "key_contact_email",
                    "verification_status",
                    "last_modified_at", 
                    "last_synced_at",
                    "key_contact_name", 
                    "key_contact_position",
                ] 
            )
        else:
            st.success("🎉 Everything is up to date! Modify records in Airtable to see them here.")

        # --- PERSISTENT LOGS AT BOTTOM ---
        if "last_sync_results" in st.session_state:
            res = st.session_state["last_sync_results"]
            st.divider()
            failures = res.get("failures", 0)
            skipped = res.get("skipped", 0)
            blocked = res.get("blocked", 0)
            deferred_text = f" / {skipped} Deferred" if skipped else ""
            blocked_text = f" / {blocked} Blocked" if blocked else ""
            with st.expander(
                f"📋 Last Sync Log ({res['timestamp']}) - {res['count']} Successes / {failures} Failures{blocked_text}{deferred_text}",
                expanded=True,
            ):
                # ── Top-line counters ───────────────────────────────────
                details = res.get("details") or []
                from collections import Counter as _C
                op_counts = _C()
                for item in details:
                    if item.get("status") == "Success":
                        op_counts[item.get("op", "Sync")] += 1
                    elif item.get("status") == "Blocked":
                        op_counts["Blocked"] += 1
                    else:
                        op_counts["Failed"] += 1

                cols = st.columns(max(len(op_counts), 1))
                for col, (op_name, cnt) in zip(cols, op_counts.most_common()):
                    col.metric(op_name, cnt)

                if blocked:
                    blocked_deleted = res.get("blocked_deleted", 0)
                    st.caption(
                        f"🛡️ Email Verification: {blocked} leads blocked "
                        f"({blocked_deleted} removed from Instantly)"
                    )

                fc = res.get("failure_counts") or {}
                if fc:
                    with st.expander("⚠️ Failure breakdown by category", expanded=False):
                        for k, v in sorted(fc.items(), key=lambda kv: kv[1], reverse=True):
                            samples = ", ".join((res.get("failure_samples") or {}).get(k, [])[:5])
                            st.write(f"- **{k}**: {v} (sample rows: {samples})")
                        st.caption(
                            "`nan_json` = NaN/Infinity in payload · "
                            "`rate_limited` = Instantly 429 · "
                            "`missing_config` = missing Instantly key or campaign creation failed"
                        )

                # ── Per-lead summary table ──────────────────────────────
                if details:
                    table_rows = []
                    for item in details:
                        if item.get("status") == "Success":
                            op = item.get("op", "Sync")
                        elif item.get("status") == "Blocked":
                            op = "Blocked"
                        else:
                            op = "Failed"
                        new_id = item.get("new_instantly_id")
                        table_rows.append({
                            "op": op,
                            "email": item.get("email") or "—",
                            "company": item.get("company_name") or "—",
                            "industry": item.get("industry") or "—",
                            "city": item.get("city") or "—",
                            "verification": item.get("verification_status") or "—",
                            "campaign": item.get("campaign_name") or "—",
                            "instantly_id": (new_id[:8] + "…") if isinstance(new_id, str) and len(new_id) > 8 else (new_id or "—"),
                            "error": item.get("error") or "",
                        })

                    import pandas as _pd
                    df_log = _pd.DataFrame(table_rows)

                    # Optional op filter
                    all_ops = sorted(df_log["op"].dropna().unique().tolist())
                    selected_ops = st.multiselect(
                        "Filter by operation",
                        options=all_ops,
                        default=all_ops,
                        key="sync_log_op_filter",
                    )
                    if selected_ops:
                        df_log = df_log[df_log["op"].isin(selected_ops)]

                    st.dataframe(
                        df_log,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "op": st.column_config.TextColumn("Op", width="small"),
                            "email": st.column_config.TextColumn("Email"),
                            "company": st.column_config.TextColumn("Company"),
                            "industry": st.column_config.TextColumn("Industry", width="small"),
                            "city": st.column_config.TextColumn("City", width="small"),
                            "verification": st.column_config.TextColumn("MV", width="small"),
                            "campaign": st.column_config.TextColumn("Campaign"),
                            "instantly_id": st.column_config.TextColumn("Instantly ID", width="small"),
                            "error": st.column_config.TextColumn("Error"),
                        },
                    )

                if st.button("🧹 Clear Logs"):
                    del st.session_state["last_sync_results"]
                    st.rerun()

    # --- TAB 3: CAMPAIGN COMPOSER ---
    with tab_camp:
        campaign_composer_ui.render(
            backend,
            secrets,
            active_mode=active_mode,
            debug_mode=debug_mode,
        )


if __name__ == "__main__":
    main()
